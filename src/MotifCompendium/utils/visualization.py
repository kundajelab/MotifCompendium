import base64
import io
import os

from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import pandas as pd

import MotifCompendium.utils.plotting as utils_plotting


####################
# PUBLIC FUNCTIONS #
####################
def motif_collection_html(
    motif_groups: dict[str, list[utils_plotting.LogoPlottingInput]], html_out: str
) -> None:
    """Creates an HTML file displaying groups of motifs.

    Creates a two panel HTML file displaying motifs grouped into sections. Motifs will be
    displayed in the order that they are listed. The HTML file is produced using jinja2
    and motif_collection_template.html.

    Args:
        motif_groups: A dictionary from group names to groups, where each group is
          specified as a list of LogoPlottingInput objects.
        html_out: The path to save the html file.
    """
    if not html_out.endswith(".html"):
        html_out += ".html"
    # Keep track of which plots come from which groups to reorder them later
    all_motifs = []
    group_to_motif_idx = dict()
    start = 0
    for group_name, group in motif_groups.items():
        end = start + len(group)
        group_to_motif_idx[group_name] = (start, end)
        start = end
        all_motifs += group
    # Plot
    all_motifs = utils_plotting.plot_many_motif_logos(all_motifs)
    # Redefine motif_groups using updated LogoPlottingInput objects
    for group_name, (start, end) in group_to_motif_idx.items():
        motif_groups[group_name] = all_motifs[start:end]
    # Create Jinja2 environment
    current_dir = os.path.dirname(os.path.abspath(__file__))
    env = Environment(loader=FileSystemLoader(current_dir))
    # Load HTML template
    template = env.get_template("motif_collection_template.html")
    # Render HTML with data
    rendered_html = template.render(data=motif_groups, sorted=sorted)
    # Write HTML to file
    with open(html_out, "w") as f:
        f.write(rendered_html)


def table_html(
    table: pd.DataFrame, image_column: list[bool], html_out: str, editable: bool
) -> None:
    """Creates an HTML file displaying the values in a pd.DataFrame.

    Creates a table HTML file of the values in a pd.DataFrame. Some of the columns can
    be UTF-8 encoded images but must be specified as such with the image_column variable.
    The HTML file is produced using jinja2 and table_template.html.

    Args:
        table: A table to display table.
        image_column: A list of booleans corresponding to each column in the table. True
          indicates that the corresponding column contains images and False indicates
          that the corresponding column contain values that will be displayed as text.
        html_out: The path to save he html file.
        editable: A boolean of whether or not the text values in the table can be edited.
    """
    # Check inputs
    if not isinstance(table, pd.DataFrame):
        raise TypeError("table must be a pd.DataFrame")
    if not html_out.endswith(".html"):
        html_out += ".html"
    # Add index column
    table.insert(0, "index", table.index)
    image_column.insert(0, False)
    # Prepare data for rendering
    columns = table.columns.tolist()
    rows = table.to_dict(orient="records")
    # Create Jinja2 environment
    current_dir = os.path.dirname(os.path.abspath(__file__))
    env = Environment(loader=FileSystemLoader(current_dir))
    # Load HTML template
    template = env.get_template("table_template.html")
    # Render HTML with data
    rendered_html = template.render(
        columns=columns, rows=rows, image_column=image_column, editable=editable
    )
    # Write HTML to file
    with open(html_out, "w") as f:
        f.write(rendered_html)


def df_to_xlsx(table: pd.DataFrame, image_column: list[bool], xlsx_out: str) -> None:
    """Creates an Excel (.xlsx) file displaying the values in a pd.DataFrame.

    Creates a .xlsx file of the values in a pd.DataFrame. Some of the columns can be
    UTF-8 encoded images but must be specified as such with the image_column variable.

    Args:
        table: A pd.DataFrame to export to .xlsx format.
        image_column: A list of booleans corresponding to each column in the table. True
          indicates that the corresponding column contains images and False indicates
          that the corresponding column contain values that will be displayed as text.
        xlsx_out: The path to save the xlsx file.
    """
    # Check inputs
    if not isinstance(table, pd.DataFrame):
        raise TypeError("table must be a pd.DataFrame")
    if not xlsx_out.endswith(".xlsx"):
        xlsx_out += ".xlsx"
    columns = table.columns.tolist()
    if len(image_column) != len(columns):
        raise ValueError("image_column length must match number of DataFrame columns")
    # Create file
    wb = Workbook()
    ws = wb.active
    ws.title = "motif summary"
    # Set width for image columns to 64.07 (≈970 pixels)
    for c_idx, is_img in enumerate(image_column, start=1):
        if is_img:
            ws.column_dimensions[get_column_letter(c_idx)].width = 64.07
    # Write header row
    for c_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=c_idx, value=str(col_name))
    # Write data rows
    for r_idx, (_, row) in enumerate(table.iterrows(), start=2):
        for c_idx, col_name in enumerate(columns, start=1):
            val = row[col_name]
            if image_column[c_idx - 1]:
                ws.add_image(
                    XLImage(io.BytesIO(base64.b64decode(val))),
                    anchor=ws.cell(row=r_idx, column=c_idx).coordinate,
                )
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)
    # Set height for all non-header rows to 127.50 (≈340 pixels)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 127.50
    # Save workbook
    wb.save(xlsx_out)
